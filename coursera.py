import argparse
from datetime import datetime
import json
from random import choice
import re

from bs4 import BeautifulSoup as soup
from openpyxl import Workbook
import requests


COURSES_URL = 'https://www.coursera.org/sitemap~www~courses.xml'
NUM_COURSES = 20


def get_courses_list():
    courses_list = []
    response = requests.get(COURSES_URL).content
    courses_soup = soup(response, 'lxml')
    courses_tags_list = courses_soup.find_all('loc')
    courses_number = 0
    while courses_number < NUM_COURSES and courses_tags_list:
        course_tag = choice(courses_tags_list)
        courses_list.append(course_tag.get_text())
        courses_tags_list.remove(course_tag)
        courses_number += 1
    return courses_list


def get_course_info(course_url):
    response = requests.get(course_url).content
    course_soup = soup(response, 'html.parser')
    course_title = course_soup.find('title').get_text()
    course_language = course_soup.find('div',
                                       class_='rc-Language'
                                       ).get_text().split(',')[0]
    course_start_date = get_course_dates(course_soup)['start_date']
    course_weeks = get_course_dates(course_soup)['weeks']
    course_rating = get_course_rating(course_soup)
    return {'title': course_title,
            'language': course_language,
            'start_date': course_start_date,
            'weeks': course_weeks,
            'rating': course_rating
            }


def get_course_rating(course_soup):
    rating_string = course_soup.find('div',
                                     class_="ratings-text bt3-hidden-xs")
    if rating_string is not None:
        return float(re.findall('[0-5][\.1-9]*', rating_string.get_text())[0])
    else:
        return 0


def get_course_dates(course_soup):
    json_string = course_soup.find('script', type="application/ld+json")
    if json_string is not None:
        json_string = json_string.get_text()
        info = json.loads(json_string)
        start_date = info['hasCourseInstance'][0]['startDate']
        end_date = info['hasCourseInstance'][0]['endDate']
        return {'start_date': start_date,
                'weeks': get_weeks_between_dates(start_date, end_date)
                }
    else:
        return {'start_date': 'not identified',
                'weeks': 'not identified'
                }


def get_weeks_between_dates(start_date, end_date):
    d1 = datetime.strptime(start_date, "%Y-%m-%d")
    d2 = datetime.strptime(end_date, "%Y-%m-%d")
    return abs((d2 - d1).days) // 7


def output_courses_info_to_xlsx(filepath):
    courses_list = get_courses_list()
    myworkbook = Workbook()
    courses_sheet = myworkbook.active
    courses_sheet['A1'] = 'Title'
    courses_sheet['B1'] = 'Language'
    courses_sheet['C1'] = 'Starting date'
    courses_sheet['D1'] = 'Number of weeks'
    courses_sheet['E1'] = 'Rated out of 5'
    course_number = 0
    for row in range(2, NUM_COURSES + 2):
        courses_sheet.cell(
            row=row, column=1).value = get_course_info(
            courses_list[course_number])['title']
        courses_sheet.cell(
            row=row, column=2).value = get_course_info(
            courses_list[course_number])['language']
        courses_sheet.cell(
            row=row, column=3).value = get_course_info(
            courses_list[course_number])['start_date']
        courses_sheet.cell(
            row=row, column=4).value = get_course_info(
            courses_list[course_number])['weeks']
        courses_sheet.cell(
            row=row, column=5).value = get_course_info(
            courses_list[course_number])['rating']
        course_number += 1
    myworkbook.save(filepath)


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('filepath', help='path to a new .xlsx file')
    args = parser.parse_args()
    output_courses_info_to_xlsx(args.filepath)
